#!/usr/bin/env python3
"""Extract the untreated Scheffler et al. zygote trajectories into a tidy CSV."""

import csv
from pathlib import Path

from openpyxl import load_workbook


HERE = Path(__file__).resolve().parent
SOURCE = HERE / "Source Data.xlsx"
OUTPUT = HERE / "scheffler_2021_control_zygote_trajectories.csv"


def read_block(ws, header_row, end_row):
    labels = [ws.cell(header_row, col).value for col in range(3, 56)]
    result = {}
    for row in range(header_row + 1, end_row + 1):
        time_h = ws.cell(row, 2).value
        if not isinstance(time_h, (int, float)):
            continue
        for offset, label in enumerate(labels, start=3):
            value = ws.cell(row, offset).value
            if label and isinstance(value, (int, float)):
                embryo = int(str(label).split()[-1])
                result[(embryo, float(time_h))] = float(value)
    return result


def main():
    # Normal mode is intentional: read-only worksheets re-scan XML on each
    # random cell access and are extremely slow for this 118-sheet workbook.
    workbook = load_workbook(SOURCE, read_only=False, data_only=True)

    distances = workbook["Figure 1b"]
    male_distance = read_block(distances, 4, 45)
    female_distance = read_block(distances, 48, 89)

    male_volume_sheet = workbook["Figure S1b"]
    female_volume_sheet = workbook["Figure S1c"]
    male_volume = read_block(male_volume_sheet, 4, 49)
    female_volume = read_block(female_volume_sheet, 4, 49)

    durations = {}
    duration_sheet = workbook["Figure S1m"]
    for row in range(4, duration_sheet.max_row + 1):
        label = duration_sheet.cell(row, 2).value
        value = duration_sheet.cell(row, 3).value
        if label and isinstance(value, (int, float)):
            durations[int(str(label).split()[-1])] = float(value)

    fieldnames = [
        "embryo_id",
        "time_h",
        "normalized_time_tau",
        "migration_duration_h",
        "male_to_center_um",
        "female_to_center_um",
        "nearer_to_center_um",
        "farther_to_center_um",
        "distance_sum_um",
        "distance_difference_um",
        "male_relative_volume",
        "female_relative_volume",
        "volume_sum",
        "volume_difference",
    ]

    rows = []
    common_keys = sorted(set(male_distance) & set(female_distance))
    for embryo, time_h in common_keys:
        duration = durations.get(embryo)
        if duration is None or time_h < 0 or time_h > duration:
            continue
        male_d = male_distance[(embryo, time_h)]
        female_d = female_distance[(embryo, time_h)]
        male_v = male_volume.get((embryo, time_h))
        female_v = female_volume.get((embryo, time_h))
        row = {
            "embryo_id": f"Scheffler2021_Z{embryo:02d}",
            "time_h": time_h,
            "normalized_time_tau": time_h / duration,
            "migration_duration_h": duration,
            "male_to_center_um": male_d,
            "female_to_center_um": female_d,
            "nearer_to_center_um": min(male_d, female_d),
            "farther_to_center_um": max(male_d, female_d),
            "distance_sum_um": male_d + female_d,
            "distance_difference_um": abs(male_d - female_d),
            "male_relative_volume": male_v,
            "female_relative_volume": female_v,
            "volume_sum": male_v + female_v if male_v is not None and female_v is not None else None,
            "volume_difference": abs(male_v - female_v) if male_v is not None and female_v is not None else None,
        }
        rows.append(row)

    with OUTPUT.open("w", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    embryo_count = len({row["embryo_id"] for row in rows})
    print(f"Wrote {len(rows)} frame rows across {embryo_count} embryos to {OUTPUT}")
    if embryo_count != 53:
        raise RuntimeError(f"Expected 53 embryos, found {embryo_count}")


if __name__ == "__main__":
    main()
